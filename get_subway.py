#!/usr/bin/env python3
#coding:utf-8

'''

https://subway.simba.taobao.com/report/getMarketAnalysis.htm?bidwordstr=%5B%22%E5%A6%88%E5%A6%88%E5%A4%96%E5%A5%97%22%5D&startDate=2017-10-03&endDate=2017-10-09

bidwordstr=["妈妈外套"]
startDate=2017-10-03
endDate=2017-10-09

sla=json&isAjaxRequest=true&token=bd540d6b&_referer=%2Ftools%2Finsight%2Fqueryresult%3Fkws%3D%25E5%25A6%2588%25E5%25A6%2588%25E5%25A4%2596%25E5%25A5%2597

Host: subway.simba.taobao.com
User-Agent: Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:56.0) Gecko/20100101 Firefox/56.0
Accept: application/json, text/javascript, */*; q=0.01
Accept-Language: zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3
Accept-Encoding: gzip, deflate, br
X-Requested-With: XMLHttpRequest
Content-Type: application/x-www-form-urlencoded; charset=UTF-8
Referer: https://subway.simba.taobao.com/
Content-Length: 153
Cookie: thw=cn; isg=AmhorqDipdd_-olT2U64QT21OF-6OdPDK7Ek-iKMIeN2fRDnzaOUKvv7Cy5z; cna=x/liEhW8J3ACATypkBfR+VvJ; t=dca9e7652d1d5b37cabeecc08b620441; cookie2=63a00a6becdf6bc54abc78835fa5cb5e; v=0; _tb_token_=5653eeee13161; um=6AF5B463492A874DE0E37EC7103B7DD4CB7652D09EBC914AEDCA26BDFB835115CCBEFCD4ACEDCB9DCD43AD3E795C914CCF409962B90BD530F9216C0A980CC896; uc1="cookie15=W5iHLLyFOGW7aA%3D%3D"; uc3=sg2=VTrqGeai3w1QnzTG4C5aPkPyvf1w2ywRVuYveMRyU6k%3D&nk2=tOyQLPbXozy5stC4&id2=UNQ%2F19SFosWR&vt3=F8dBzLBEfoz2nbyG5RI%3D&lg2=WqG3DMC9VAQiUQ%3D%3D; existShop=MTUwNzU5MzI5Nw%3D%3D; uss=WqPyBXaK12Mi8eQFu5RRtmraZ7drXUiuqx3kK8FLP3EMRHbfUci4KTjM; lgc=%5Cu5DE6%5Cu53F3%5Cu9022%5Cu6E90%5Cu670D%5Cu9970; tracknick=%5Cu5DE6%5Cu53F3%5Cu9022%5Cu6E90%5Cu670D%5Cu9970; sg=%E9%A5%B002; mt=ci=1_1; cookie1=B0OsnZwBMl%2FjhSiIsyCo7ozokPqXefx2axGQK1v7dtI%3D; unb=348151420; skt=b1e65c79b05d103d; publishItemObj=Ng%3D%3D; _cc_=V32FPkk%2Fhw%3D%3D; tg=0; _l_g_=Ug%3D%3D; _nk_=%5Cu5DE6%5Cu53F3%5Cu9022%5Cu6E90%5Cu670D%5Cu9970; cookie17=UNQ%2F19SFosWR; _m_h5_tk=2e9140b21f43d863881cc0c58b6fe15e_1507595643896; _m_h5_tk_enc=b729a0f70f2bff8ce58e3db7f145e110; x=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0; whl=-1%260%260%261507597664965; JSESSIONID=C1AE756CEC38E8CA8881FC8D07F15809; linezing_session=MsVvQFJ33fKuilWbgfHwvqyY_1507593478712HdrB_1; hng=CN%7Czh-CN%7CCNY%7C156; apush538637cbd2f43b0a2e2df52a0ad01eb7=%7B%22ts%22%3A1507598623213%2C%22parentId%22%3A1507597680378%7D
DNT: 1
Connection: keep-alive




bidwordstr	外套妈妈
date	2017-10-03
impression	147697  展现指数
impressionRate	null
click	5268       点击指数
price	1195462
ctr	333           点击率 ， 除以100
competition	2410  竞争度
cvr	74            点击转化率，除以100
avgPrice	190   市场平均出价，除以100

直通车搜索列表，flag=0 是，你是不是想找
https://subway.simba.taobao.com/bidword/tool/adgroup/getSuggestWord.htm?adgroupId=798039979&word=%E5%A6%88%E5%A6%88&productId=101001005

adgroupId:798039979
word:妈妈
productId:101001005

sla:json
isAjaxRequest:true
token:e77ae2dd
_referer:/campaigns/standards/adgroups/items/detail?tab=bidword&campaignId=40133048&adGroupId=798039979

'''
from selenium import webdriver
import time , json ,os ,urllib ,datetime
import requests
import openpyxl



def openbrowser_login():
    driver=webdriver.Chrome('../chromedriver')
    driver.get('http://zhitongche.taobao.com/')
    while(True):
        if(len(driver.window_handles)>1):
           print('检测到页面跳转！')
           #切换到新页面
           driver.switch_to.window(driver.window_handles[1]);
           time.sleep(3)
           #重新找开新页面，以便获取新页面的cookies
           driver.get(driver.current_url)
           time.sleep(5)
           break;
        else:
           time.sleep(2)
    cookie = [item["name"] + "=" + item["value"] for item in driver.get_cookies()]
    cookiestr=';'.join(item for item in cookie)
    with open('./work/cookie', 'wt') as fs:
        fs.write(cookiestr)
    print('登陆cookie已经保存!')
    try:
        driver.quit()
    except Exception as e:
        pass
    return cookiestr

def check_login(cookiestr):
    print('开始登陆验证!')
    url = 'https://i.taobao.com/my_taobao.htm'
    headers = {
        'Host': 'i.taobao.com',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:43.0) Gecko/20100101 Firefox/43.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3',
        'Referer': 'https://www.taobao.com',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Connection': 'Keep-Alive',
        'Cookie': cookiestr,
        'Cache-Control': 'max-age=0',
    }

    try:
        req = requests.get(url, headers=headers, )
        if (req.url == url):
            print('登陆验证通过!')
            return True
    except Exception as e:
        print(e)
    print('登陆验证失败!请重新登陆!')
    return False

def check_subway(cookiestr):
    print('开始淘宝直通车验证!')
    url='http://subway.simba.taobao.com/bpenv/getLoginUserInfo.htm'
    headers= {
            'Host':'subway.simba.taobao.com',
            'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:43.0) Gecko/20100101 Firefox/43.0',
            'Accept':'application/json, text/javascript, */*; q=0.01',
            'Accept-Language':'zh-CN,zh;q=0.8',
            'Connection' : 'Keep-Alive',
            'Cookie' : cookiestr,
            'Origin':'http://subway.simba.taobao.com',
            'Cache-Control':'max-age=0',
            'X-Requested-With':'XMLHttpRequest'
        }
    postdata={'_referer':'/tools/insight/index'}
    try:
        req = requests.post(url, headers=headers ,data=postdata, )
        parse=json.loads(req.text)
        if(parse['code']=='200'):
            print('淘宝直通车验证通过!您当前以<'+parse['result']['nickName']+'>登陆')
            return parse['result']['token']
    except Exception as e:
        print(e)
    print('淘宝直通车验证失败!请重新登陆!')
    return False


class  subwayquery:
    url='http://subway.simba.taobao.com/report/getNetworkPerspective.htm'

    def __init__(self,token,cookiestr):
       self.data= {
           'sla':'json',
           'isAjaxRequest':'true',
           'token':token,
           '_referer':'/tools/insight/queryresult?tab=tabs-region&start=&end=&kws=',
           }
       self.token=token
       self.startDate=(datetime.datetime.now()-datetime.timedelta(days=7)).strftime('%Y-%m-%d')
       self.endDate=(datetime.datetime.now()-datetime.timedelta(days=1)).strftime('%Y-%m-%d')
       self.headers= {
            'Host':'subway.simba.taobao.com',
            'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:43.0) Gecko/20100101 Firefox/43.0',
            'Accept':'application/json, text/javascript, */*; q=0.01',
            'Accept-Language':'zh-CN,zh;q=0.8',
            'Connection' : 'Keep-Alive',
            'Cookie' : cookiestr,
            'Origin':'https://subway.simba.taobao.com',
            'Cache-Control':'max-age=0',
            'X-Requested-With':'XMLHttpRequest',
            'Referer':'https://subway.simba.taobao.com/'
        }

    def get_keyword_by_input(self,keywords=[]):
        '''

        :param keywords: keywords为空是获取默认热词的数据,不为空则是获取输入关键的相关词数据
        :return:
        外层增加了一个列表,数据结构为
        [[关键词1数据],[关键词2数据]]
        写入文件或应用过滤规则时,先遍历一遍

        '''
        parse_data=[]
        data = {
            'sla': 'json',
            'isAjaxRequest': 'true',
            'token': self.token,
            'referer': '/campaigns/standards/adgroups/items/detail?tab=bidword&campaignId=40133048&adGroupId=798039979',
        }
        if keywords:
            i = 3
            for keyword in keywords:
                try:
                    print('获取第{:d}行名为{}相关联的热词'.format(i , keyword))
                    tourl="https://subway.simba.taobao.com/bidword/tool/adgroup/relative.htm?pageSize=800&wordPackage=16&adGroupId=798039979&queryWord={}&orderBy=3&productId=101001005".format(urllib.parse.quote(keyword))
                    re_data = self.get_parse_data(tourl,data)
                    print('{}共有{:d}个相关词'.format(keyword, len(re_data)))
                    parse_data.extend(re_data)
                    i+=1
                    #time.sleep(1)
                except:
                    pass
        else:
            print('获取默认关键词数据')
            tourl = 'https://subway.simba.taobao.com/bidword/tool/adgroup/recommend.htm?wordPackage=16&adGroupId=798039979&orderBy=3&platForm=1&pageSize=800&productId=101001005'
            re_data = self.get_parse_data(tourl,data)
            print('共获取到{:d}个相关词'.format(len(re_data)))
            parse_data.extend(re_data)
        return parse_data

    def save_data(self , data ,fs = './work/key.xlsx'):
        if os.path.exists(fs):
            wb = openpyxl.load_workbook(fs)
            ws = wb['key']
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'key'
            title = ['关键词', '相关度', '平均出价', '展现指数', '点击率', '点击转化率', '竞争度', '词包']
            ws.append([])
            ws.append(title)
        maxrow = ws.max_row
        for item in data:
            maxrow += 1
            #print(maxrow, '\t', item['word'], '\t', item['averagePrice'], '\t', item['ctr'], '\t', item['cvr'])
            #保留相关度为5的词
            c1 = ws.cell(row=maxrow, column=1)
            c1.value = item['word']
            c2 = ws.cell(row=maxrow, column=2)
            c2.value = item['pertinence']
            c3 = ws.cell(row=maxrow, column=3)
            c3.value = float(item['averagePrice']) / 100 if item['averagePrice'] else 0
            c4 = ws.cell(row=maxrow, column=4)
            c4.value = item['pv']
            c5 = ws.cell(row=maxrow, column=5)
            c5.value = item['ctr']
            c6 = ws.cell(row=maxrow, column=6)
            c6.value = item['cvr']
            c7 = ws.cell(row=maxrow, column=7)
            c7.value = item['competition']
            c7 = ws.cell(row=maxrow, column=8)
            c7.value = item['wordPackage']
        wb.save(fs)
        wb.close()

    def get_filter_data(self,data,filter_data=[]):
        '''
        将源数据通或指定关键词过滤,包含filter_data中词的将被过滤掉
        filter_data默认为空列表
        '''
        if filter_data:
            remain_data = data[:]
            print('数据过滤包含关键词{}'.format(filter_data))
            print('筛选前源数据有{:d}项'.format(len(remain_data)))
            for key in data:
                for del_key in filter_data:
                    if del_key in key['word']:
                        remain_data.remove(key)
                        break
            print('筛选后源数据有{:d}项'.format(len(remain_data)))
        else:
            remain_data=data
        return remain_data


    def get_parse_data(self , url , data):
        try:
            #10次重试保证返回值
            i=0;
            while(True):
                response = requests.post(url , headers = self.headers , data = data)
                parse=json.loads(response.text)
                if(parse['code']=='200'):
                    if(parse['result']):
                        return parse['result']
                    else:
                        i+=1
                        if(i==10):
                     #       print(keyword)
                            return False
                else:
                    return False
            return False
        except Exception as e:
            print(e)
            return False


    def query(self,keyword,perspectiveType):
        tourl=self.url+'?bidwordstr='+urllib.parse.quote(keyword)+'&startDate='+self.startDate+'&endDate='+self.endDate+'&perspectiveType='+perspectiveType
        data= {
           'sla':'json',
           'isAjaxRequest':'true',
           'token':self.token,
           '_referer':'/tools/insight/queryresult?kws='+urllib.parse.quote(keyword)+'&tab=tabs-region&start=&end=',
           }
        parse_data = self.get_parse_data(tourl , data)
        return parse_data

class read_write:
    def __init__(self,token,cookiestr):
        self.subway=subwayquery(token,cookiestr)
        #self.threadpool=threadpool.ThreadPool(32)
        self.centeralignment=openpyxl.styles.Alignment(horizontal='center')
        self.percentage_format=openpyxl.styles.numbers.FORMAT_PERCENTAGE_00
        self.number00_format=openpyxl.styles.numbers.FORMAT_NUMBER_00
        self.number_format=openpyxl.styles.numbers.FORMAT_NUMBER

    def write_sheet(self,sheet,data,row,base):
        cell1=sheet.cell(row=row,column=base+1)
        cell1.value=float(data['impressionRate'])/10000
        cell1.number_format=self.percentage_format
        cell2=sheet.cell(row=row,column=base+2)
        cell2.value=int(data['impression'])
        cell2.number_format=self.number_format
        cell3=sheet.cell(row=row,column=base+3)
        cell3.value=int(data['click'])
        cell3.number_format=self.number_format
        cell4=sheet.cell(row=row,column=base+4)
        cell4.value=float(data['ctr'])/10000
        cell4.number_format=self.percentage_format
        cell5=sheet.cell(row=row,column=base+5)
        cell5.value=float(data['cvr'])/10000
        cell5.number_format=self.percentage_format
        cell6=sheet.cell(row=row,column=base+6)
        cell6.value=float(data['avgPrice'])/100
        cell6.number_format=self.number00_format
        cell7=sheet.cell(row=row,column=base+7)
        cell7.value=int(data['competition'])
        cell7.number_format=self.number_format

    def query_write(self,sheet,row):
        keyword=sheet.cell(row=row,column=1).value
        '''
        #站内站外流量解析
        try:
            parse1=self.subway.query(keyword,'1')
            print(parse1)
            if(parse1!=False):
                self.write_sheet(sheet,parse1['result'][0],row,18)       #4
                self.write_sheet(sheet,parse1['result'][1],row,25)      #11
        except Exception as e:
            print("%s,%d,%s,%s" % (keyword,row,e,parse1['result']))
            pass
        
        '''
        # PC端及无线端解析
        try:
            parse2=self.subway.query(keyword,'2')
            #print(parse2)
            if(parse2!=False):
                self.write_sheet(sheet,parse2['result'][0],row,7)      #18
                self.write_sheet(sheet,parse2['result'][1],row,14)      #25
        except Exception as e:
            print("%s,%d,%s,%s" % (keyword,row,e,parse2))
            pass

    def try_save(self,wb,filename):
        try:
            print('正在保存文件：'+filename)
            wb.save(filename)
        except  Exception as e:
            wb.save(filename+'_copy')

    def start(self):
        files=os.listdir(os.getcwd()+'/work')
        for filename in files:
            print('正在读取文件：'+filename)
            wb=openpyxl.load_workbook('work//'+filename)
            sheets=wb.sheetnames
            for sheetname in sheets:
                print('正在读取表:'+sheetname)
                sheet=wb[sheetname]
                sheet.merge_cells('A1:E1')
                sheet.merge_cells('H1:N1')
                sheet.merge_cells('O1:U1')
                #sheet.merge_cells('S1:Y1')
                #sheet.merge_cells('Z1:AF1')
                #sheet['E1'].value='淘宝站内'
                #sheet['L1'].value='淘宝站外'
                sheet['E1'].value='推荐热词'
                sheet['H1'].value='计算机设备'
                sheet['O1'].value='移动设备'
                sheet['H1'].alignment=sheet['O1'].alignment=self.centeralignment
                sheet['H2'].value=sheet['O2'].value='展现占比'
                sheet['I2'].value=sheet['P2'].value='展现指数'
                sheet['J2'].value=sheet['Q2'].value='点击指数'
                sheet['K2'].value=sheet['R2'].value='点击率'
                sheet['L2'].value=sheet['S2'].value='点击转化率'
                sheet['M2'].value=sheet['T2'].value='市场均价'
                sheet['N2'].value=sheet['U2'].value='竞争度'
                max_row=sheet.max_row
                for row in range(3,max_row+1):
                    #start = time.time()
                    self.query_write(sheet,row)
                    #end = time.time()
                    #print('第%d行执行完毕，用时%.1f秒.' % (row, (end - start)))

                '''
                for i in range(0,(max_row+2997)//3000):
                    start=time.time()
                    arguments=list()
                    for row in range(3+i*3000,min(3002+i*3000,max_row)+1):
                        arguments.append(([sheet,row],()))
                    requests=threadpool.makeRequests(self.query_write,arguments)
                    [self.threadpool.putRequest(req) for req in requests]
                    self.threadpool.wait()
                    end=time.time()
                    print('第%d-%d行执行完毕，用时%.1f秒.'%(3+i*3000,min(3002+i*3000,max_row),(end-start)))
                '''
            self.try_save(wb,filename)



def main():
    if (os.path.exists('./work/cookie')):
        print('检测到cookie文件！将使用cookie登陆！')
        with open('./work/cookie', 'r') as fp : cookiestr = fp.read()
    else:
        cookiestr = openbrowser_login()
    while (True):
        if (check_login(cookiestr)):
            token = check_subway(cookiestr)
            if (token != False):
                break;
        cookiestr = openbrowser_login()
    que = subwayquery(token,cookiestr)
    keys = []
    fs = './work/key-1.xlsx'
    try:
        wb = openpyxl.load_workbook(fs)
        ws = wb.active

        for row in range(3,ws.max_row):
            keys.append(ws.cell(row=row,column=1).value)
        print(keys)
        wb.close()
    except:
        pass
    start_time=time.time()
    data = que.get_keyword_by_input(keys)

    que.save_data(que.get_filter_data(data) , fs)
    print("共用时{:.2f}秒".format(time.time() - start_time))

    #根据关键词表查每个词数据透视，再将每个关键词数据写入xlsx中
    #save_excel = read_write(token,cookiestr)
    #save_excel.start()

if __name__=="__main__":
    main()


